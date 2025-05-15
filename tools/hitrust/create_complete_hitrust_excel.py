#!/usr/bin/env python3
"""
Create Complete HITRUST CSF v11 Excel Framework for CISO Assistant
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

def create_complete_hitrust_excel():
    print("Creating HITRUST CSF v11 Excel file...")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create required sheets
    sheets_to_create = [
        'library_meta',
        'framework_meta',
        'framework_content',
        'reference_controls_meta',
        'reference_controls_content',
        'scores_definition_meta',
        'scores_definition_content',
        'implementation_groups_definition_meta',
        'implementation_groups_definition_content',
        'risk_matrix_meta',
        'risk_matrix_content'
    ]
    
    for sheet_name in sheets_to_create:
        wb.create_sheet(sheet_name)
    
    # Populate library_meta
    ws_lib_meta = wb['library_meta']
    library_meta_data = [
        ["type", "library"],
        ["urn", "urn:intuitem:risk:library:hitrust-csf-v11"],
        ["locale", "en"],
        ["ref_id", "HITRUST-CSF-v11"],
        ["name", "HITRUST CSF v11"],
        ["description", "HITRUST Common Security Framework (CSF) version 11"],
        ["copyright", "© 2024 HITRUST Alliance"],
        ["version", "11"],
        ["publication_date", str(date.today())],
        ["provider", "HITRUST Alliance"],
        ["packager", "intuitem"]
    ]
    
    for row in library_meta_data:
        ws_lib_meta.append(row)
    
    # Populate framework_meta
    ws_fw_meta = wb['framework_meta']
    framework_meta_data = [
        ["type", "framework"],
        ["urn", "urn:intuitem:risk:framework:hitrust-csf-v11"],
        ["ref_id", "HITRUST-CSF-v11"],
        ["name", "HITRUST CSF v11"],
        ["description", "HITRUST Common Security Framework version 11"],
        ["implementation_groups_definition", "implementation_groups_definition"],
        ["scores_definition", "scores_definition"],
        ["base_urn", "urn:intuitem:risk:req_node:hitrust-csf-v11"]
    ]
    
    for row in framework_meta_data:
        ws_fw_meta.append(row)
    
    # Populate framework_content with complete HITRUST structure
    ws_fw_content = wb['framework_content']
    fw_headers = ["assessable", "depth", "ref_id", "name", "description", "implementation_groups", "threats", "reference_controls", "annotation", "typical_evidence", "questions"]
    ws_fw_content.append(fw_headers)
    
    # Complete HITRUST CSF v11 control structure (all categories 00-13)
    framework_content = [
        # Category 00: Information Security Management Program
        ["false", 1, "00", "00 - Information Security Management Program", "Control family for establishing and operating the information security management program", "", "", "", "", "", ""],
        ["true", 2, "00.a", "00.a - Information Security Policy", "Establishment and maintenance of information security policies", "IG1,IG2,IG3", "", "reference_controls:RC-01", "These policies are aligned to the organizational business goals, objectives and requirements and explicitly define the risk tolerance of senior leadership", "Information security policy documentation, policy governance structure", ""],
        ["true", 2, "00.b", "00.b - Security Roles and Responsibilities", "Definition of information security roles and responsibilities", "IG1,IG2,IG3", "", "reference_controls:RC-01", "", "Role descriptions, organizational chart with security roles", ""],
        ["true", 2, "00.c", "00.c - Information Security in Project Management", "Addressing information security within project management", "IG2,IG3", "", "reference_controls:RC-01", "", "Project security requirements, security gates documentation", ""],
        ["true", 2, "00.d", "00.d - Segregation of Duties", "Appropriate segregation of duties to reduce risk", "IG2,IG3", "", "reference_controls:RC-01", "", "Segregation matrix, role conflict analysis", ""],
        ["true", 2, "00.e", "00.e - Contact with Authorities", "Procedures for contacting relevant authorities", "IG2,IG3", "", "reference_controls:RC-01", "", "Authority contact procedures, emergency contact lists", ""],
        ["true", 2, "00.f", "00.f - Contact with Special Interest Groups", "Maintaining contact with security groups and forums", "IG3", "", "reference_controls:RC-01", "", "Membership documentation, interaction records", ""],
        
        # Category 01: Access Control
        ["false", 1, "01", "01 - Access Control", "Control family for access management and control", "", "", "", "", "", ""],
        ["true", 2, "01.a", "01.a - Access Control Policy", "Access control policy based on business requirements", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Access control policy, access management procedures", ""],
        ["true", 2, "01.b", "01.b - Access to Networks and Network Services", "Control of access to networks and network services", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Network access controls documentation", ""],
        ["true", 2, "01.c", "01.c - User Registration and De-registration", "User account provisioning and deprovisioning", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "User provisioning procedures, account logs", ""],
        ["true", 2, "01.d", "01.d - User Access Provisioning", "Authorization for access rights", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Access request forms, approval records", ""],
        ["true", 2, "01.e", "01.e - Management of Privileged Access Rights", "Control over privileged access", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Privileged access management procedures", ""],
        ["true", 2, "01.f", "01.f - Review of User Access Rights", "Periodic review of access rights", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Access review reports", ""],
        ["true", 2, "01.g", "01.g - Removal or Adjustment of Access Rights", "Timely adjustment of access rights", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Access revocation logs", ""],
        ["true", 2, "01.h", "01.h - Use of Secret Authentication Information", "Protection of authentication information", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Password policy, authentication guidelines", ""],
        ["true", 2, "01.i", "01.i - Secure Log-on Procedures", "Secure authentication procedures", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Authentication logs, MFA documentation", ""],
        ["true", 2, "01.j", "01.j - Password Management System", "System for managing passwords", "IG2,IG3", "", "reference_controls:RC-02", "", "Password management system documentation", ""],
        ["true", 2, "01.k", "01.k - Use of Privileged Utility Programs", "Control of system utilities", "IG2,IG3", "", "reference_controls:RC-02", "", "Utility access logs", ""],
        ["true", 2, "01.l", "01.l - Access Control to Program Source Code", "Protection of source code", "IG2,IG3", "", "reference_controls:RC-02", "", "Code repository access controls", ""],
        ["true", 2, "01.m", "01.m - Information Access Restriction", "Restriction based on access control policy", "IG1,IG2,IG3", "", "reference_controls:RC-02", "", "Data access matrix", ""],
        ["true", 2, "01.n", "01.n - Sensitive System Isolation", "Isolation of sensitive systems", "IG3", "", "reference_controls:RC-02", "", "System isolation documentation", ""],
        
        # Category 02: Human Resources Security
        ["false", 1, "02", "02 - Human Resources Security", "Control family for personnel security", "", "", "", "", "", ""],
        ["true", 2, "02.a", "02.a - Screening", "Background verification checks", "IG1,IG2,IG3", "", "reference_controls:RC-03", "", "Background check procedures", ""],
        ["true", 2, "02.b", "02.b - Terms and Conditions of Employment", "Employment agreements including security", "IG1,IG2,IG3", "", "reference_controls:RC-03", "", "Employment contracts with security clauses", ""],
        ["true", 2, "02.c", "02.c - Management Responsibilities", "Management's role in security", "IG2,IG3", "", "reference_controls:RC-03", "", "Management security responsibilities", ""],
        ["true", 2, "02.d", "02.d - Information Security Awareness, Education, and Training", "Security training programs", "IG1,IG2,IG3", "", "reference_controls:RC-03", "", "Training materials and records", ""],
        ["true", 2, "02.e", "02.e - Disciplinary Process", "Process for security violations", "IG1,IG2,IG3", "", "reference_controls:RC-03", "", "Disciplinary procedures", ""],
        ["true", 2, "02.f", "02.f - Termination or Change of Employment", "Procedures for employment changes", "IG1,IG2,IG3", "", "reference_controls:RC-03", "", "Termination checklists", ""],
        
        # Category 03: Risk Management
        ["false", 1, "03", "03 - Risk Management", "Control family for risk management processes", "", "", "", "", "", ""],
        ["true", 2, "03.a", "03.a - Risk Management Program", "Formal risk management program", "IG1,IG2,IG3", "", "reference_controls:RC-04", "", "Risk management framework documentation", ""],
        ["true", 2, "03.b", "03.b - Risk Assessment", "Regular risk assessments", "IG1,IG2,IG3", "", "reference_controls:RC-04", "", "Risk assessment reports", ""],
        ["true", 2, "03.c", "03.c - Risk Treatment", "Risk treatment plans", "IG1,IG2,IG3", "", "reference_controls:RC-04", "", "Risk treatment plans and tracking", ""],
        ["true", 2, "03.d", "03.d - Risk Acceptance", "Formal risk acceptance process", "IG2,IG3", "", "reference_controls:RC-04", "", "Risk acceptance documentation", ""],
        ["true", 2, "03.e", "03.e - Risk Monitoring and Review", "Ongoing monitoring of risks", "IG2,IG3", "", "reference_controls:RC-04", "", "Risk monitoring reports", ""],
        
        # Category 04: Asset Management
        ["false", 1, "04", "04 - Asset Management", "Control family for asset management", "", "", "", "", "", ""],
        ["true", 2, "04.a", "04.a - Inventory of Assets", "Comprehensive asset inventory", "IG1,IG2,IG3", "", "reference_controls:RC-05", "", "Asset inventory database", ""],
        ["true", 2, "04.b", "04.b - Ownership of Assets", "Assignment of asset ownership", "IG1,IG2,IG3", "", "reference_controls:RC-05", "", "Asset ownership registry", ""],
        ["true", 2, "04.c", "04.c - Acceptable Use of Assets", "Rules for acceptable asset use", "IG1,IG2,IG3", "", "reference_controls:RC-05", "", "Acceptable use policy", ""],
        ["true", 2, "04.d", "04.d - Return of Assets", "Asset return upon termination", "IG1,IG2,IG3", "", "reference_controls:RC-05", "", "Asset return procedures", ""],
        ["true", 2, "04.e", "04.e - Classification of Information", "Information classification scheme", "IG2,IG3", "", "reference_controls:RC-05", "", "Classification policy and labels", ""],
        ["true", 2, "04.f", "04.f - Labeling of Information", "Information labeling procedures", "IG2,IG3", "", "reference_controls:RC-05", "", "Labeling standards", ""],
        ["true", 2, "04.g", "04.g - Handling of Assets", "Secure handling procedures", "IG2,IG3", "", "reference_controls:RC-05", "", "Asset handling procedures", ""],
        
        # Category 05: Physical and Environmental Security
        ["false", 1, "05", "05 - Physical and Environmental Security", "Control family for physical security", "", "", "", "", "", ""],
        ["true", 2, "05.a", "05.a - Physical Security Perimeter", "Physical boundaries for security", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Facility floor plans with perimeters", ""],
        ["true", 2, "05.b", "05.b - Physical Entry Controls", "Physical access control systems", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Access logs, badge system documentation", ""],
        ["true", 2, "05.c", "05.c - Securing Offices, Rooms, and Facilities", "Physical security of work areas", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Physical security assessments", ""],
        ["true", 2, "05.d", "05.d - Protecting Against External and Environmental Threats", "Environmental protection measures", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Environmental controls documentation", ""],
        ["true", 2, "05.e", "05.e - Working in Secure Areas", "Procedures for secure areas", "IG2,IG3", "", "reference_controls:RC-06", "", "Secure area procedures", ""],
        ["true", 2, "05.f", "05.f - Delivery and Loading Areas", "Security of delivery areas", "IG2,IG3", "", "reference_controls:RC-06", "", "Loading dock procedures", ""],
        ["true", 2, "05.g", "05.g - Siting and Protection of Equipment", "Equipment placement and protection", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Equipment placement guidelines", ""],
        ["true", 2, "05.h", "05.h - Supporting Utilities", "Utility infrastructure protection", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Utility redundancy documentation", ""],
        ["true", 2, "05.i", "05.i - Cabling Security", "Protection of cabling", "IG2,IG3", "", "reference_controls:RC-06", "", "Cabling diagrams and protection measures", ""],
        ["true", 2, "05.j", "05.j - Equipment Maintenance", "Maintenance procedures", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Maintenance logs", ""],
        ["true", 2, "05.k", "05.k - Removal of Assets", "Control of asset removal", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Asset removal authorization logs", ""],
        ["true", 2, "05.l", "05.l - Security of Equipment and Assets Off-premises", "Off-site security measures", "IG2,IG3", "", "reference_controls:RC-06", "", "Off-site security procedures", ""],
        ["true", 2, "05.m", "05.m - Secure Disposal or Reuse of Equipment", "Secure disposal procedures", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Disposal certificates", ""],
        ["true", 2, "05.n", "05.n - Unattended User Equipment", "Protection of unattended equipment", "IG1,IG2,IG3", "", "reference_controls:RC-06", "", "Screen lock policies", ""],
        ["true", 2, "05.o", "05.o - Clear Desk and Clear Screen Policy", "Workspace security", "IG2,IG3", "", "reference_controls:RC-06", "", "Clean desk policy", ""],
        
        # Category 06: Communications and Operations Management
        ["false", 1, "06", "06 - Communications and Operations Management", "Control family for operational security", "", "", "", "", "", ""],
        ["true", 2, "06.a", "06.a - Documented Operating Procedures", "Documentation of operating procedures", "IG1,IG2,IG3", "", "reference_controls:RC-07", "", "Operating procedures, runbooks", ""],
        ["true", 2, "06.b", "06.b - Change Management", "Control of changes", "IG1,IG2,IG3", "", "reference_controls:RC-07", "", "Change management records", ""],
        ["true", 2, "06.c", "06.c - Capacity Management", "Monitoring and managing capacity", "IG2,IG3", "", "reference_controls:RC-07", "", "Capacity reports", ""],
        ["true", 2, "06.d", "06.d - Separation of Development, Testing, and Operational Environments", "Environment separation", "IG2,IG3", "", "reference_controls:RC-07", "", "Environment architecture diagrams", ""],
        ["true", 2, "06.e", "06.e - Protection from Malware", "Anti-malware controls", "IG1,IG2,IG3", "", "reference_controls:RC-07", "", "Antivirus logs and configuration", ""],
        ["true", 2, "06.f", "06.f - Information Backup", "Backup procedures", "IG1,IG2,IG3", "", "reference_controls:RC-07", "", "Backup logs and test reports", ""],
        ["true", 2, "06.g", "06.g - Event Logging", "Security event logging", "IG1,IG2,IG3", "", "reference_controls:RC-08", "", "Log configuration and retention", ""],
        ["true", 2, "06.h", "06.h - Protection of Log Information", "Security of audit logs", "IG2,IG3", "", "reference_controls:RC-08", "", "Log protection measures", ""],
        ["true", 2, "06.i", "06.i - Administrator and Operator Logs", "Logging of administrative actions", "IG2,IG3", "", "reference_controls:RC-08", "", "Admin activity logs", ""],
        ["true", 2, "06.j", "06.j - Clock Synchronization", "System time synchronization", "IG2,IG3", "", "reference_controls:RC-07", "", "NTP configuration", ""],
        ["true", 2, "06.k", "06.k - Installation of Software on Operational Systems", "Control of software installation", "IG2,IG3", "", "reference_controls:RC-07", "", "Software installation procedures", ""],
        ["true", 2, "06.l", "06.l - Network Controls", "Network security management", "IG1,IG2,IG3", "", "reference_controls:RC-09", "", "Network diagrams and configurations", ""],
        ["true", 2, "06.m", "06.m - Segregation in Networks", "Network segmentation", "IG2,IG3", "", "reference_controls:RC-09", "", "Network segmentation documentation", ""],
        ["true", 2, "06.n", "06.n - Network Connection Control", "Control of network connections", "IG2,IG3", "", "reference_controls:RC-09", "", "Firewall rules and ACLs", ""],
        ["true", 2, "06.o", "06.o - Network Routing Control", "Control of network routing", "IG3", "", "reference_controls:RC-09", "", "Routing tables and policies", ""],
        ["true", 2, "06.p", "06.p - Formal Transfer Policies", "Information transfer policies", "IG2,IG3", "", "reference_controls:RC-09", "", "Data transfer procedures", ""],
        ["true", 2, "06.q", "06.q - Procedures and Controls for Media in Transit", "Protection of media in transit", "IG2,IG3", "", "reference_controls:RC-09", "", "Media transport procedures", ""],
        ["true", 2, "06.r", "06.r - Electronic Messaging", "Security of electronic messaging", "IG2,IG3", "", "reference_controls:RC-09", "", "Email security configuration", ""],
        ["true", 2, "06.s", "06.s - Business Information Systems", "Security within business systems", "IG2,IG3", "", "reference_controls:RC-07", "", "Business system security documentation", ""],
        ["true", 2, "06.t", "06.t - Publicly Available Information", "Control of public information", "IG2,IG3", "", "reference_controls:RC-07", "", "Public information procedures", ""],
        ["true", 2, "06.u", "06.u - Electronic Commerce", "Security of e-commerce", "IG3", "", "reference_controls:RC-09", "", "E-commerce security controls", ""],
        ["true", 2, "06.v", "06.v - On-Line Transactions", "Security of online transactions", "IG3", "", "reference_controls:RC-09", "", "Transaction security logs", ""],
        
        # Category 07: System Development and Maintenance
        ["false", 1, "07", "07 - System Development and Maintenance", "Control family for secure development", "", "", "", "", "", ""],
        ["true", 2, "07.a", "07.a - Security Requirements Analysis and Specification", "Security in requirements", "IG2,IG3", "", "reference_controls:RC-10", "", "Security requirements documentation", ""],
        ["true", 2, "07.b", "07.b - Secure Development Policy", "Policy for secure development", "IG2,IG3", "", "reference_controls:RC-10", "", "Secure SDLC policy", ""],
        ["true", 2, "07.c", "07.c - Change Control Procedures", "Control of system changes", "IG2,IG3", "", "reference_controls:RC-10", "", "Change control documentation", ""],
        ["true", 2, "07.d", "07.d - Technical Review of Applications After Operating Platform Changes", "Technical reviews", "IG3", "", "reference_controls:RC-10", "", "Technical review reports", ""],
        ["true", 2, "07.e", "07.e - Restrictions on Changes to Software Packages", "Control of packaged software", "IG2,IG3", "", "reference_controls:RC-10", "", "Software change controls", ""],
        ["true", 2, "07.f", "07.f - System Security Testing", "Security testing procedures", "IG2,IG3", "", "reference_controls:RC-10", "", "Security test reports", ""],
        ["true", 2, "07.g", "07.g - Outsourced Development", "Security in outsourced development", "IG3", "", "reference_controls:RC-10", "", "Outsourcing security requirements", ""],
        ["true", 2, "07.h", "07.h - Protection of Application Services on Public Networks", "Public application security", "IG3", "", "reference_controls:RC-10", "", "Public-facing application assessments", ""],
        ["true", 2, "07.i", "07.i - Securing Application Services", "Application service security", "IG2,IG3", "", "reference_controls:RC-10", "", "Application security architecture", ""],
        ["true", 2, "07.j", "07.j - Technical Vulnerability Management", "Vulnerability management process", "IG1,IG2,IG3", "", "reference_controls:RC-10", "", "Vulnerability scan reports", ""],
        
        # Category 08: Information Security Incident Management
        ["false", 1, "08", "08 - Information Security Incident Management", "Control family for incident response", "", "", "", "", "", ""],
        ["true", 2, "08.a", "08.a - Incident Response Planning", "Incident response plan development", "IG1,IG2,IG3", "", "reference_controls:RC-11", "", "Incident response plan", ""],
        ["true", 2, "08.b", "08.b - Incident Response Training", "Training for incident response", "IG2,IG3", "", "reference_controls:RC-11", "", "Training records", ""],
        ["true", 2, "08.c", "08.c - Incident Response Testing", "Testing of incident response", "IG2,IG3", "", "reference_controls:RC-11", "", "Test reports", ""],
        ["true", 2, "08.d", "08.d - Incident Response Handling", "Incident handling procedures", "IG1,IG2,IG3", "", "reference_controls:RC-11", "", "Incident handling playbooks", ""],
        ["true", 2, "08.e", "08.e - Incident Response Monitoring", "Monitoring for incidents", "IG2,IG3", "", "reference_controls:RC-11", "", "Monitoring configuration", ""],
        ["true", 2, "08.f", "08.f - Incident Response Reporting", "Incident reporting procedures", "IG1,IG2,IG3", "", "reference_controls:RC-11", "", "Incident reporting forms", ""],
        ["true", 2, "08.g", "08.g - Collection of Evidence", "Evidence collection procedures", "IG2,IG3", "", "reference_controls:RC-11", "", "Evidence handling procedures", ""],
        ["true", 2, "08.h", "08.h - Learning from Information Security Incidents", "Post-incident improvement", "IG2,IG3", "", "reference_controls:RC-11", "", "Lessons learned documentation", ""],
        
        # Category 09: Business Continuity Management
        ["false", 1, "09", "09 - Business Continuity Management", "Control family for business continuity", "", "", "", "", "", ""],
        ["true", 2, "09.a", "09.a - Business Continuity Management Policy", "BCM policy establishment", "IG2,IG3", "", "reference_controls:RC-12", "", "BCM policy documentation", ""],
        ["true", 2, "09.b", "09.b - Implementing Business Continuity", "BCM implementation procedures", "IG2,IG3", "", "reference_controls:RC-12", "", "BCM implementation plans", ""],
        ["true", 2, "09.c", "09.c - Business Continuity and Risk Assessment", "Risk assessment for BCM", "IG2,IG3", "", "reference_controls:RC-12", "", "Business impact analysis", ""],
        ["true", 2, "09.d", "09.d - Developing and Implementing Continuity Plans", "Continuity plan development", "IG2,IG3", "", "reference_controls:RC-12", "", "Business continuity plans", ""],
        ["true", 2, "09.e", "09.e - Business Continuity Plan Testing", "Testing of BC plans", "IG2,IG3", "", "reference_controls:RC-12", "", "BC test reports", ""],
        ["true", 2, "09.f", "09.f - Business Continuity Plan Maintenance", "Maintenance of BC plans", "IG2,IG3", "", "reference_controls:RC-12", "", "BC plan update logs", ""],
        ["true", 2, "09.g", "09.g - Information System Recovery and Restoration", "IS recovery procedures", "IG2,IG3", "", "reference_controls:RC-12", "", "Recovery playbooks", ""],
        ["true", 2, "09.h", "09.h - Information System Backup", "Backup strategy for recovery", "IG1,IG2,IG3", "", "reference_controls:RC-12", "", "Backup procedures and tests", ""],
        
        # Category 10: Compliance
        ["false", 1, "10", "10 - Compliance", "Control family for regulatory compliance", "", "", "", "", "", ""],
        ["true", 2, "10.a", "10.a - Identification of Applicable Legislation", "Legal requirements identification", "IG1,IG2,IG3", "", "reference_controls:RC-13", "", "Legal register", ""],
        ["true", 2, "10.b", "10.b - Intellectual Property Rights", "IPR compliance procedures", "IG2,IG3", "", "reference_controls:RC-13", "", "IPR compliance documentation", ""],
        ["true", 2, "10.c", "10.c - Protection of Organizational Records", "Records protection procedures", "IG1,IG2,IG3", "", "reference_controls:RC-13", "", "Records retention schedule", ""],
        ["true", 2, "10.d", "10.d - Data Protection and Privacy", "Privacy compliance measures", "IG1,IG2,IG3", "", "reference_controls:RC-13", "", "Privacy impact assessments", ""],
        ["true", 2, "10.e", "10.e - Prevention of Misuse of Information Processing Facilities", "Acceptable use enforcement", "IG2,IG3", "", "reference_controls:RC-13", "", "Usage monitoring reports", ""],
        ["true", 2, "10.f", "10.f - Regulation of Cryptographic Controls", "Cryptography compliance", "IG2,IG3", "", "reference_controls:RC-13", "", "Crypto compliance documentation", ""],
        ["true", 2, "10.g", "10.g - Compliance with Security Policies and Standards", "Security compliance checking", "IG2,IG3", "", "reference_controls:RC-13", "", "Compliance assessment reports", ""],
        ["true", 2, "10.h", "10.h - Technical Compliance Checking", "Technical compliance validation", "IG2,IG3", "", "reference_controls:RC-13", "", "Technical compliance scan reports", ""],
        ["true", 2, "10.i", "10.i - Information Systems Audit Controls", "IS audit control measures", "IG2,IG3", "", "reference_controls:RC-13", "", "Audit scope and procedures", ""],
        ["true", 2, "10.j", "10.j - Protection of Information Systems Audit Tools", "Audit tool protection", "IG3", "", "reference_controls:RC-13", "", "Audit tool access logs", ""],
        
        # Category 11: Mobile Security
        ["false", 1, "11", "11 - Mobile Security", "Control family for mobile device security", "", "", "", "", "", ""],
        ["true", 2, "11.a", "11.a - Mobile Device Policy", "Policy for mobile device use", "IG1,IG2,IG3", "", "reference_controls:RC-14", "", "Mobile device policy", ""],
        ["true", 2, "11.b", "11.b - Mobile Code", "Control of mobile code", "IG2,IG3", "", "reference_controls:RC-14", "", "Mobile code policy", ""],
        ["true", 2, "11.c", "11.c - Teleworking", "Security for remote work", "IG2,IG3", "", "reference_controls:RC-14", "", "Telework security procedures", ""],
        ["true", 2, "11.d", "11.d - Mobile Device Management", "MDM implementation", "IG2,IG3", "", "reference_controls:RC-14", "", "MDM configuration", ""],
        ["true", 2, "11.e", "11.e - BYOD Security", "Bring Your Own Device controls", "IG2,IG3", "", "reference_controls:RC-14", "", "BYOD policy and controls", ""],
        
        # Category 12: Supplier Relationships
        ["false", 1, "12", "12 - Supplier Relationships", "Control family for third-party management", "", "", "", "", "", ""],
        ["true", 2, "12.a", "12.a - Information Security in Supplier Relationships", "Security requirements for suppliers", "IG2,IG3", "", "reference_controls:RC-15", "", "Supplier security requirements", ""],
        ["true", 2, "12.b", "12.b - Supplier Agreement Security", "Security in supplier agreements", "IG2,IG3", "", "reference_controls:RC-15", "", "Supplier contracts with security clauses", ""],
        ["true", 2, "12.c", "12.c - Supplier Service Delivery", "Monitoring supplier services", "IG2,IG3", "", "reference_controls:RC-15", "", "Supplier performance reports", ""],
        ["true", 2, "12.d", "12.d - Supplier Relationship Monitoring", "Ongoing supplier monitoring", "IG2,IG3", "", "reference_controls:RC-15", "", "Supplier audit reports", ""],
        ["true", 2, "12.e", "12.e - Supplier Transition Management", "Managing supplier changes", "IG3", "", "reference_controls:RC-15", "", "Supplier transition plans", ""],
        
        # Category 13: Cloud Security
        ["false", 1, "13", "13 - Cloud Security", "Control family for cloud computing security", "", "", "", "", "", ""],
        ["true", 2, "13.a", "13.a - Cloud Security Architecture", "Secure cloud architecture", "IG2,IG3", "", "reference_controls:RC-16", "", "Cloud architecture documentation", ""],
        ["true", 2, "13.b", "13.b - Cloud Data Protection", "Protection of cloud data", "IG2,IG3", "", "reference_controls:RC-16", "", "Cloud data encryption documentation", ""],
        ["true", 2, "13.c", "13.c - Cloud Identity and Access Management", "Cloud IAM controls", "IG2,IG3", "", "reference_controls:RC-16", "", "Cloud IAM configuration", ""],
        ["true", 2, "13.d", "13.d - Cloud Monitoring and Logging", "Cloud security monitoring", "IG2,IG3", "", "reference_controls:RC-16", "", "Cloud monitoring configuration", ""],
        ["true", 2, "13.e", "13.e - Cloud Incident Response", "Cloud-specific incident response", "IG2,IG3", "", "reference_controls:RC-16", "", "Cloud incident response procedures", ""],
        ["true", 2, "13.f", "13.f - Cloud Compliance", "Cloud compliance management", "IG2,IG3", "", "reference_controls:RC-16", "", "Cloud compliance assessments", ""],
        ["true", 2, "13.g", "13.g - Cloud Exit Planning", "Cloud service exit strategy", "IG3", "", "reference_controls:RC-16", "", "Cloud exit plans", ""]
    ]
    
    for row in framework_content:
        ws_fw_content.append(row)
    
    # Populate reference_controls_meta
    ws_rc_meta = wb['reference_controls_meta']
    rc_meta_data = [
        ["type", "reference_controls"],
        ["name", "reference_controls"],
        ["description", "Reference controls for HITRUST CSF v11"],
        ["base_urn", "urn:intuitem:risk:reference_control:hitrust-csf-v11"]
    ]
    
    for row in rc_meta_data:
        ws_rc_meta.append(row)
    
    # Populate reference_controls_content
    ws_rc_content = wb['reference_controls_content']
    rc_headers = ["ref_id", "name", "description", "category"]
    ws_rc_content.append(rc_headers)
    
    reference_controls = [
        ["RC-01", "Policy Controls", "Organizational policies and procedures", "policy"],
        ["RC-02", "Access Controls", "Access management and authentication", "technical"],
        ["RC-03", "Human Resources Controls", "Personnel security controls", "process"],
        ["RC-04", "Risk Management Controls", "Risk assessment and treatment", "process"],
        ["RC-05", "Asset Management Controls", "Asset inventory and classification", "process"],
        ["RC-06", "Physical Security Controls", "Physical and environmental security", "physical"],
        ["RC-07", "Operations Controls", "Operational procedures and controls", "technical"],
        ["RC-08", "Logging and Monitoring Controls", "Security event logging and monitoring", "technical"],
        ["RC-09", "Network Security Controls", "Network protection measures", "technical"],
        ["RC-10", "System Development Controls", "Secure development lifecycle", "process"],
        ["RC-11", "Incident Response Controls", "Incident management procedures", "process"],
        ["RC-12", "Business Continuity Controls", "Continuity planning and testing", "process"],
        ["RC-13", "Compliance Controls", "Legal and regulatory compliance", "process"],
        ["RC-14", "Mobile Security Controls", "Mobile device protection", "technical"],
        ["RC-15", "Supplier Management Controls", "Third-party risk management", "process"],
        ["RC-16", "Cloud Security Controls", "Cloud-specific security measures", "technical"]
    ]
    
    for row in reference_controls:
        ws_rc_content.append(row)
    
    # Populate scores_definition_meta
    ws_scores_meta = wb['scores_definition_meta']
    scores_meta_data = [
        ["type", "scores"],
        ["name", "scores_definition"],
        ["description", "HITRUST maturity levels"]
    ]
    
    for row in scores_meta_data:
        ws_scores_meta.append(row)
    
    # Populate scores_definition_content
    ws_scores_content = wb['scores_definition_content']
    scores_headers = ["score", "name", "description"]
    ws_scores_content.append(scores_headers)
    
    scores_data = [
        [1, "Policy", "Control requirement is documented in policies"],
        [2, "Procedure", "Procedures are documented and implemented"],
        [3, "Implemented", "Control is fully implemented"],
        [4, "Tested", "Control effectiveness is tested"],
        [5, "Managed", "Control is managed and continuously improved"]
    ]
    
    for row in scores_data:
        ws_scores_content.append(row)
    
    # Populate implementation_groups_definition_meta
    ws_ig_meta = wb['implementation_groups_definition_meta']
    ig_meta_data = [
        ["type", "implementation_groups"],
        ["name", "implementation_groups_definition"],
        ["description", "HITRUST implementation group definitions"]
    ]
    
    for row in ig_meta_data:
        ws_ig_meta.append(row)
    
    # Populate implementation_groups_definition_content
    ws_ig_content = wb['implementation_groups_definition_content']
    ig_headers = ["ref_id", "name", "description"]
    ws_ig_content.append(ig_headers)
    
    ig_data = [
        ["IG1", "Foundational", "Basic implementation for smaller organizations"],
        ["IG2", "Managed", "Enhanced controls for mid-size organizations"],
        ["IG3", "Advanced", "Comprehensive controls for large organizations"]
    ]
    
    for row in ig_data:
        ws_ig_content.append(row)
    
    # Populate risk_matrix_meta
    ws_rm_meta = wb['risk_matrix_meta']
    rm_meta_data = [
        ["type", "risk_matrix"],
        ["urn", "urn:intuitem:risk:risk_matrix:hitrust-csf-v11"],
        ["ref_id", "HITRUST-CSF-v11-RM"],
        ["name", "HITRUST CSF v11 Risk Matrix"],
        ["description", "Risk assessment matrix for HITRUST CSF v11"]
    ]
    
    for row in rm_meta_data:
        ws_rm_meta.append(row)
    
    # Populate risk_matrix_content
    ws_rm_content = wb['risk_matrix_content']
    rm_headers = ["type", "id", "color", "abbreviation", "name", "description", "grid0", "grid1", "grid2", "grid3", "grid4"]
    ws_rm_content.append(rm_headers)
    
    # Apply colors to risk matrix cells
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Risk matrix data
    rm_data = [
        # Probability levels
        ["probability", 0, "", "VL", "Very Low", "Probability < 10%", 0, 0, 1, 2, 3],
        ["probability", 1, "", "L", "Low", "Probability 10-35%", 0, 1, 2, 3, 4],
        ["probability", 2, "", "M", "Medium", "Probability 35-65%", 1, 2, 3, 3, 4],
        ["probability", 3, "", "H", "High", "Probability 65-90%", 2, 3, 3, 4, 4],
        ["probability", 4, "", "VH", "Very High", "Probability > 90%", 3, 4, 4, 4, 4],
        # Impact levels
        ["impact", 0, "", "VL", "Very Low", "Minimal impact"],
        ["impact", 1, "", "L", "Low", "Minor impact"],
        ["impact", 2, "", "M", "Medium", "Moderate impact"],
        ["impact", 3, "", "H", "High", "Major impact"],
        ["impact", 4, "", "VH", "Very High", "Severe impact"],
        # Risk levels
        ["risk", 0, "", "VL", "Very Low", "Minimal risk"],
        ["risk", 1, "", "L", "Low", "Low risk"],
        ["risk", 2, "", "M", "Medium", "Medium risk"],
        ["risk", 3, "", "H", "High", "High risk"],
        ["risk", 4, "", "VH", "Very High", "Critical risk"]
    ]
    
    for row_idx, row in enumerate(rm_data):
        ws_rm_content.append(row)
        
        # Apply colors to risk levels
        if row[0] == "risk":
            row_num = ws_rm_content.max_row
            if row[1] == 0:
                ws_rm_content.cell(row=row_num, column=3).fill = green_fill
            elif row[1] in [1, 2]:
                ws_rm_content.cell(row=row_num, column=3).fill = yellow_fill
            elif row[1] == 3:
                ws_rm_content.cell(row=row_num, column=3).fill = orange_fill
            elif row[1] == 4:
                ws_rm_content.cell(row=row_num, column=3).fill = red_fill
    
    # Apply colors to probability grid cells
    colors_map = {0: green_fill, 1: yellow_fill, 2: yellow_fill, 3: orange_fill, 4: red_fill}
    
    for row_idx in range(2, 7):  # Probability rows
        for col_idx in range(7, 12):  # Grid columns
            cell = ws_rm_content.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.fill = colors_map.get(cell.value, PatternFill())
    
    # Format headers
    for ws in wb.worksheets:
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.font = Font(bold=True)
    
    # Save the file
    file_path = "/Users/ashishthirunagari/Documents/GitHub/ciso-assistant-community/tools/hitrust/HITRUST_CSF_v11_Complete.xlsx"
    wb.save(file_path)
    
    print(f"✅ Created complete HITRUST CSF v11 Excel file: {file_path}")
    print(f"   • 14 control categories (00-13)")
    print(f"   • {len(framework_content)} controls")
    print(f"   • {len(reference_controls)} reference controls")
    print(f"   • Complete risk matrix")
    print(f"   • Implementation groups")
    
    return file_path

if __name__ == "__main__":
    create_complete_hitrust_excel()
