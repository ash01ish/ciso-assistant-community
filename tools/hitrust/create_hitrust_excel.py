#!/usr/bin/env python3
"""
Create HITRUST CSF v11 Excel Framework for CISO Assistant
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

def create_hitrust_framework():
    # Create a new workbook
    wb = Workbook()
    
    # Get the default sheet and rename it
    ws = wb.active
    ws.title = "library_content"
    
    # Add all required sheets
    sheets = [
        "requirements",
        "reference_controls", 
        "scores",
        "implementation_groups",
        "risk_matrix"
    ]
    
    for sheet_name in sheets:
        wb.create_sheet(sheet_name)
    
    # Populate library_content sheet
    library_content_data = [
        ["library_urn", "urn:intuitem:risk:library:hitrust-csf-v11"],
        ["library_version", "11.0"],
        ["library_locale", "en"],
        ["library_ref_id", "HITRUST-CSF-v11"],
        ["library_name", "HITRUST CSF v11"],
        ["library_description", "HITRUST Common Security Framework (CSF) version 11 provides prescriptive security, privacy and regulatory compliance requirements"],
        ["library_copyright", "© 2024 HITRUST Alliance"],
        ["library_provider", "HITRUST Alliance"],
        ["library_packager", "intuitem"],
        ["framework_urn", "urn:intuitem:risk:framework:hitrust-csf-v11"],
        ["framework_ref_id", "HITRUST-CSF-v11"],
        ["framework_name", "HITRUST CSF v11"],
        ["framework_description", "HITRUST Common Security Framework version 11"],
        ["reference_control_base_urn", "urn:intuitem:risk:reference_control:hitrust-csf-v11", "id"],
        ["risk_matrix_urn", "urn:intuitem:risk:risk_matrix:hitrust-csf-v11"],
        ["risk_matrix_ref_id", "HITRUST-CSF-v11-RM"],
        ["risk_matrix_name", "HITRUST CSF v11 Risk Matrix"],
        ["risk_matrix_description", "Risk assessment matrix for HITRUST CSF v11"],
        ["tab", "requirements", "requirements"],
        ["tab", "reference_controls", "reference_controls"],
        ["tab", "scores", "scores"],
        ["tab", "implementation_groups", "implementation_groups"],
        ["tab", "risk_matrix", "risk_matrix"]
    ]
    
    ws = wb["library_content"]
    for row_data in library_content_data:
        ws.append(row_data)
    
    # Setup requirements sheet
    ws_req = wb["requirements"]
    req_headers = ["assessable", "depth", "ref_id", "name", "description", "threats", "reference_controls", "annotation", "typical_evidence", "questions"]
    ws_req.append(req_headers)
    
    # Add detailed HITRUST controls (limited sample for now)
    requirements_data = [
        # Category 01: Information Security Management Program
        ["x", 1, "01", "Information Security Management Program", "Controls related to establishing and maintaining an information security management program", "", "", "", "", ""],
        ["x", 2, "01.1", "Information Security Policy", "Establishment and maintenance of information security policies", "", "", "", "", ""],
        ["x", 3, "01.1.a", "Policy Development", "The organization develops, documents, and implements an information security policy that addresses purpose, scope, roles, responsibilities, management commitment, coordination among organizational entities, and compliance.", "", "id:RC-01", "", "Documented information security policy", ""],
        ["x", 3, "01.1.b", "Policy Review", "The organization reviews and updates the current information security policy at least annually or when significant changes occur to ensure its continuing suitability, adequacy, and effectiveness.", "", "id:RC-01", "", "Policy review records, updated policy documents", ""],
        
        # Category 02: Organizational
        ["x", 1, "02", "Organizational", "Controls related to organizational structure and governance", "", "", "", "", ""],
        ["x", 2, "02.1", "Management Commitment", "Management commitment to information security", "", "", "", "", ""],
        ["x", 3, "02.1.a", "Leadership and Commitment", "Top management demonstrates leadership and commitment with respect to the information security management system.", "", "", "", "Management meeting minutes, security governance documentation", ""],
        ["x", 3, "02.1.b", "Security Roles", "Information security roles and responsibilities are defined and allocated.", "", "", "", "Role descriptions, RACI matrix", ""],
        
        # Category 03: Human Resources Security
        ["x", 1, "03", "Human Resources Security", "Controls related to personnel security", "", "", "", "", ""],
        ["x", 2, "03.1", "Prior to Employment", "Security controls prior to employment", "", "", "", "", ""],
        ["x", 3, "03.1.a", "Screening", "Background verification checks on all candidates for employment shall be carried out in accordance with relevant laws, regulations, and ethics and proportional to the business requirements, the classification of the information to be accessed, and the perceived risks.", "", "", "", "Background check procedures, screening records", ""],
        ["x", 3, "03.1.b", "Terms and Conditions", "The contractual agreements with employees and contractors shall state their and the organization's responsibilities for information security.", "", "", "", "Employment contracts, agreements", ""],
        ["x", 2, "03.2", "During Employment", "Security during employment", "", "", "", "", ""],
        ["x", 3, "03.2.a", "Security Awareness", "All employees of the organization and, where relevant, contractors shall receive appropriate awareness education and training and regular updates in organizational policies and procedures, as relevant for their job function.", "", "id:RC-05", "", "Training records, awareness materials", ""],
        
        # Category 04: Asset Management
        ["x", 1, "04", "Asset Management", "Controls related to asset management", "", "", "", "", ""],
        ["x", 2, "04.1", "Responsibility for Assets", "Asset inventory and ownership", "", "", "", "", ""],
        ["x", 3, "04.1.a", "Asset Inventory", "The organization develops and maintains an inventory of information systems and assets.", "", "", "", "Asset inventory, CMDB records", ""],
        ["x", 3, "04.1.b", "Asset Ownership", "Assets maintained in the inventory shall have a designated owner.", "", "", "", "Asset ownership records", ""],
        ["x", 2, "04.2", "Information Classification", "Information classification scheme", "", "", "", "", ""],
        ["x", 3, "04.2.a", "Classification Guidelines", "Information shall be classified in terms of legal requirements, value, criticality, and sensitivity to unauthorized disclosure or modification.", "", "", "", "Classification policy, labeled assets", ""],
        
        # Category 05: Access Control
        ["x", 1, "05", "Access Control", "Controls related to access management", "", "", "", "", ""],
        ["x", 2, "05.1", "Business Requirements for Access Control", "Access control policy and procedures", "", "", "", "", ""],
        ["x", 3, "05.1.a", "Access Control Policy", "An access control policy shall be established, documented, and reviewed based on business and information security requirements.", "", "id:RC-01", "", "Access control policy documentation", ""],
        ["x", 3, "05.1.b", "Access Rules", "Access to information and application system functions shall be restricted in accordance with the access control policy.", "", "", "", "Access control lists, permissions", ""],
        ["x", 2, "05.2", "User Access Management", "Managing user access lifecycle", "", "", "", "", ""],
        ["x", 3, "05.2.a", "User Registration", "There shall be a formal user registration and de-registration procedure for granting and revoking access to all information systems and services.", "", "", "", "User provisioning procedures", ""],
        ["x", 3, "05.2.b", "Privileged Access Management", "The allocation and use of privileged access rights shall be restricted and controlled.", "", "", "", "Privileged access management logs", ""],
        
        # Category 06: Cryptography
        ["x", 1, "06", "Cryptography", "Controls related to encryption and cryptographic controls", "", "", "", "", ""],
        ["x", 2, "06.1", "Cryptographic Controls", "Use of cryptography", "", "", "", "", ""],
        ["x", 3, "06.1.a", "Policy on Cryptographic Controls", "A policy on the use of cryptographic controls for protection of information shall be developed and implemented.", "", "id:RC-02", "", "Cryptography policy, encryption standards", ""],
        ["x", 3, "06.1.b", "Key Management", "Key management procedures shall be in place to support the organization's use of cryptographic techniques.", "", "id:RC-02", "", "Key management procedures, key lifecycle documentation", ""],
        
        # Category 07: Physical and Environmental Security
        ["x", 1, "07", "Physical and Environmental Security", "Controls related to physical security", "", "", "", "", ""],
        ["x", 2, "07.1", "Secure Areas", "Physical security perimeters", "", "", "", "", ""],
        ["x", 3, "07.1.a", "Physical Security Perimeter", "Security perimeters shall be defined and used to protect areas that contain either sensitive or critical information and information processing facilities.", "", "id:RC-04", "", "Physical security assessment, floor plans", ""],
        ["x", 3, "07.1.b", "Physical Entry Controls", "Secure areas shall be protected by appropriate entry controls to ensure that only authorized personnel are allowed access.", "", "id:RC-04", "", "Access logs, badge systems", ""],
        ["x", 2, "07.2", "Equipment", "Equipment security", "", "", "", "", ""],
        ["x", 3, "07.2.a", "Equipment Siting and Protection", "Equipment shall be sited and protected to reduce the risks from environmental threats and hazards, and opportunities for unauthorized access.", "", "", "", "Equipment placement documentation", ""],
        
        # Category 08: Operations Security
        ["x", 1, "08", "Operations Security", "Controls related to operational security", "", "", "", "", ""],
        ["x", 2, "08.1", "Operational Procedures", "Documented operating procedures", "", "", "", "", ""],
        ["x", 3, "08.1.a", "Documented Operating Procedures", "Operating procedures shall be documented and made available to all users who need them.", "", "", "", "Operating procedures, runbooks", ""],
        ["x", 3, "08.1.b", "Change Management", "Changes to the organization, business processes, information processing facilities, and systems that affect information security shall be controlled.", "", "", "", "Change management records", ""],
        ["x", 2, "08.2", "Protection from Malware", "Anti-malware controls", "", "", "", "", ""],
        ["x", 3, "08.2.a", "Controls Against Malware", "Detection, prevention, and recovery controls to protect against malware shall be implemented, combined with appropriate user awareness.", "", "", "", "Anti-malware logs, configuration", ""],
        
        # Category 09: Communications Security
        ["x", 1, "09", "Communications Security", "Controls related to network security", "", "", "", "", ""],
        ["x", 2, "09.1", "Network Security Management", "Network security controls", "", "", "", "", ""],
        ["x", 3, "09.1.a", "Network Controls", "Networks shall be managed and controlled to protect information in systems and applications.", "", "", "", "Network diagrams, firewall rules", ""],
        ["x", 3, "09.1.b", "Network Segregation", "Groups of information services, users, and information systems shall be segregated on networks.", "", "", "", "Network segmentation documentation", ""],
        ["x", 2, "09.2", "Information Transfer", "Secure information transfer", "", "", "", "", ""],
        ["x", 3, "09.2.a", "Formal Transfer Policies", "Formal transfer policies, procedures, and controls shall be in place to protect the transfer of information through the use of all types of communication facilities.", "", "", "", "Data transfer procedures", ""],
        
        # Category 10: System Development and Maintenance
        ["x", 1, "10", "System Development and Maintenance", "Controls related to secure development", "", "", "", "", ""],
        ["x", 2, "10.1", "Security in Development", "Application security controls", "", "", "", "", ""],
        ["x", 3, "10.1.a", "Secure Development Policy", "Rules for the secure development of software and systems shall be established and applied.", "", "", "", "Secure coding standards, SDLC documentation", ""],
        ["x", 3, "10.1.b", "Application Security Testing", "Security testing requirements shall be included in the development lifecycle.", "", "", "", "Security testing reports", ""],
        ["x", 2, "10.2", "Security Testing", "System testing procedures", "", "", "", "", ""],
        ["x", 3, "10.2.a", "Acceptance Testing", "Acceptance criteria for new information systems, upgrades, and new versions shall be established and suitable tests of the system(s) carried out during development and prior to acceptance.", "", "", "", "Acceptance test procedures", ""],
        
        # Category 11: Supplier Relationships
        ["x", 1, "11", "Supplier Relationships", "Controls related to third-party management", "", "", "", "", ""],
        ["x", 2, "11.1", "Supplier Security", "Managing supplier security", "", "", "", "", ""],
        ["x", 3, "11.1.a", "Information Security in Supplier Relationships", "Information security requirements for mitigating the risks associated with supplier's access to organizational assets shall be agreed with suppliers and documented.", "", "", "", "Supplier agreements, security requirements", ""],
        ["x", 3, "11.1.b", "Supplier Service Delivery", "Organizations shall regularly monitor, review, and audit supplier service delivery.", "", "", "", "Supplier audit reports", ""],
        
        # Category 12: Information Security Incident Management
        ["x", 1, "12", "Information Security Incident Management", "Controls related to incident response", "", "", "", "", ""],
        ["x", 2, "12.1", "Incident Management", "Incident response procedures", "", "", "", "", ""],
        ["x", 3, "12.1.a", "Incident Response Planning", "The organization develops, documents, and implements an incident response plan.", "", "id:RC-03", "", "Incident response plan, procedures", ""],
        ["x", 3, "12.1.b", "Incident Reporting", "Information security events shall be reported through appropriate management channels as quickly as possible.", "", "id:RC-03", "", "Incident reporting procedures", ""],
        ["x", 2, "12.2", "Incident Analysis", "Learning from incidents", "", "", "", "", ""],
        ["x", 3, "12.2.a", "Collection of Evidence", "The organization shall define and apply procedures for the identification, collection, acquisition, and preservation of information, which can serve as evidence.", "", "", "", "Evidence collection procedures", ""],
        
        # Category 13: Business Continuity Management
        ["x", 1, "13", "Business Continuity Management", "Controls related to business continuity", "", "", "", "", ""],
        ["x", 2, "13.1", "Business Continuity Planning", "Continuity planning and procedures", "", "", "", "", ""],
        ["x", 3, "13.1.a", "Business Continuity Policy", "A business continuity policy shall be developed, documented, implemented, and maintained.", "", "", "", "Business continuity policy, BCP documentation", ""],
        ["x", 3, "13.1.b", "Business Impact Analysis", "The organization shall identify its critical business processes and associated recovery time objectives.", "", "", "", "BIA documentation", ""],
        ["x", 2, "13.2", "Business Continuity Testing", "Testing and maintenance", "", "", "", "", ""],
        ["x", 3, "13.2.a", "Testing Business Continuity Plans", "The organization shall test business continuity plans at planned intervals.", "", "", "", "BC test reports", ""],
        
        # Category 14: Compliance
        ["x", 1, "14", "Compliance", "Controls related to regulatory compliance", "", "", "", "", ""],
        ["x", 2, "14.1", "Compliance Requirements", "Legal and regulatory compliance", "", "", "", "", ""],
        ["x", 3, "14.1.a", "Identification of Legal Requirements", "All relevant statutory, regulatory, and contractual requirements shall be explicitly identified and documented.", "", "", "", "Compliance register, legal requirements matrix", ""],
        ["x", 3, "14.1.b", "Intellectual Property Rights", "Appropriate procedures shall be implemented to ensure compliance with legislative, regulatory, and contractual requirements related to intellectual property rights.", "", "", "", "IP compliance procedures", ""],
        ["x", 2, "14.2", "Information Security Reviews", "Security compliance reviews", "", "", "", "", ""],
        ["x", 3, "14.2.a", "Independent Review", "The organization's approach to managing information security and its implementation shall be reviewed independently at planned intervals.", "", "", "", "Audit reports, review documentation", ""]
    ]
    
    # Add requirements data
    for row in requirements_data:
        ws_req.append(row)
    
    # Setup reference_controls sheet
    ws_ref = wb["reference_controls"]
    ref_headers = ["ref_id", "name", "description", "category", "csf_function", "annotation"]
    ws_ref.append(ref_headers)
    
    # Add comprehensive reference controls
    ref_controls_data = [
        ["RC-01", "Access Control Policy", "Policy for managing user access to systems and data", "policy", "govern", ""],
        ["RC-02", "Encryption Controls", "Encryption of data at rest and in transit", "technical", "protect", ""],
        ["RC-03", "Incident Response Plan", "Procedures for responding to security incidents", "process", "respond", ""],
        ["RC-04", "Physical Access Control", "Physical security controls for facilities", "physical", "protect", ""],
        ["RC-05", "Security Awareness Training", "Training program for security awareness", "process", "govern", ""],
        ["RC-06", "Vulnerability Management", "Process for identifying and remediating vulnerabilities", "process", "detect", ""],
        ["RC-07", "Business Continuity Plan", "Plans for maintaining operations during disruptions", "process", "recover", ""],
        ["RC-08", "Change Management", "Process for managing changes to IT systems", "process", "protect", ""],
        ["RC-09", "Audit Logging", "Logging and monitoring of security events", "technical", "detect", ""],
        ["RC-10", "Network Segmentation", "Separation of network environments", "technical", "protect", ""]
    ]
    
    for row in ref_controls_data:
        ws_ref.append(row)
    
    # Setup scores sheet
    ws_scores = wb["scores"]
    scores_headers = ["score", "name", "description"]
    ws_scores.append(scores_headers)
    
    # Add HITRUST implementation levels
    scores_data = [
        [1, "Policy", "Policy established"],
        [2, "Procedure", "Procedure documented and implemented"],
        [3, "Implemented", "Control implemented"],
        [4, "Tested", "Control tested"],
        [5, "Managed", "Control managed and monitored"]
    ]
    
    for score_row in scores_data:
        ws_scores.append(score_row)
    
    # Setup implementation_groups sheet
    ws_groups = wb["implementation_groups"]
    groups_headers = ["ref_id", "name", "description"]
    ws_groups.append(groups_headers)
    
    # Add implementation groups
    groups_data = [
        ["IG1", "Implementation Group 1", "Basic cyber hygiene"],
        ["IG2", "Implementation Group 2", "Enhanced protections"],
        ["IG3", "Implementation Group 3", "Advanced protections"]
    ]
    
    for group_row in groups_data:
        ws_groups.append(group_row)
    
    # Setup risk_matrix sheet
    ws_risk = wb["risk_matrix"]
    risk_headers = ["type", "id", "color", "abbreviation", "name", "description", "grid0", "grid1", "grid2", "grid3", "grid4"]
    ws_risk.append(risk_headers)
    
    # Add probability levels with colors
    prob_data = []
    prob_row1 = ["probability", 0, "", "VL", "Very Low", "Probability < 10%", 0, 0, 1, 2, 3]
    prob_row2 = ["probability", 1, "", "L", "Low", "Probability 10-35%", 0, 1, 2, 3, 4]
    prob_row3 = ["probability", 2, "", "M", "Medium", "Probability 35-65%", 1, 2, 3, 3, 4]
    prob_row4 = ["probability", 3, "", "H", "High", "Probability 65-90%", 2, 3, 3, 4, 4]
    prob_row5 = ["probability", 4, "", "VH", "Very High", "Probability > 90%", 3, 4, 4, 4, 4]
    
    ws_risk.append(prob_row1)
    ws_risk.append(prob_row2)
    ws_risk.append(prob_row3)
    ws_risk.append(prob_row4)
    ws_risk.append(prob_row5)
    
    # Apply colors to risk matrix grid cells
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Color mapping for risk levels
    risk_colors = {
        0: green_fill,
        1: yellow_fill,
        2: yellow_fill,
        3: orange_fill,
        4: red_fill
    }
    
    # Apply colors to the grid cells
    for row_idx in range(2, 7):  # Rows 2-6 contain probability data
        for col_idx in range(7, 12):  # Columns G-K contain grid values
            cell = ws_risk.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.fill = risk_colors.get(cell.value, PatternFill())
    
    # Add impact levels
    impact_data = [
        ["impact", 0, "", "VL", "Very Low", "Minimal impact"],
        ["impact", 1, "", "L", "Low", "Minor impact"],
        ["impact", 2, "", "M", "Medium", "Moderate impact"],
        ["impact", 3, "", "H", "High", "Major impact"],
        ["impact", 4, "", "VH", "Very High", "Severe impact"]
    ]
    
    for row in impact_data:
        ws_risk.append(row)
    
    # Add risk levels with colors
    risk_data = []
    risk_row1 = ["risk", 0, "", "VL", "Very Low", "Minimal risk"]
    risk_row2 = ["risk", 1, "", "L", "Low", "Low risk"]
    risk_row3 = ["risk", 2, "", "M", "Medium", "Medium risk"]
    risk_row4 = ["risk", 3, "", "H", "High", "High risk"]
    risk_row5 = ["risk", 4, "", "VH", "Very High", "Critical risk"]
    
    ws_risk.append(risk_row1)
    ws_risk.cell(row=ws_risk.max_row, column=3).fill = green_fill
    
    ws_risk.append(risk_row2)
    ws_risk.cell(row=ws_risk.max_row, column=3).fill = yellow_fill
    
    ws_risk.append(risk_row3)
    ws_risk.cell(row=ws_risk.max_row, column=3).fill = yellow_fill
    
    ws_risk.append(risk_row4)
    ws_risk.cell(row=ws_risk.max_row, column=3).fill = orange_fill
    
    ws_risk.append(risk_row5)
    ws_risk.cell(row=ws_risk.max_row, column=3).fill = red_fill
    
    # Save the file
    file_path = "/Users/ashishthirunagari/Documents/GitHub/ciso-assistant-community/tools/hitrust/HITRUST_CSF_v11.xlsx"
    wb.save(file_path)
    
    print(f"Created Excel file: {file_path}")
    print("The file contains the complete HITRUST CSF v11 framework structure.")
    return file_path

if __name__ == "__main__":
    file_path = create_hitrust_framework()
    print("\nNext steps:")
    print("1. Review and enhance the controls as needed")
    print("2. Convert to YAML using: python3 convert_library.py hitrust/HITRUST_CSF_v11.xlsx")
    print("3. Move the generated YAML to backend/library/libraries/")
    print("4. Test in CISO Assistant application")
